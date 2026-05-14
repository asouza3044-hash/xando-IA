"""
Importa CSV (gerado pelo OCR) para a aba 02_APONTAMENTO_HIST de CRONOANALISE.xlsx.

Uso:
    python importar_csv.py ocr_output\20260509_lote.csv

Regras:
- Le CSV com separador `;`
- Acrescenta linhas no final do bloco existente (sem sobrescrever)
- Mantem formatacao de borda da planilha
- Backup automatico antes de gravar: CRONOANALISE.xlsx -> CRONOANALISE_bkp_AAAAMMDD_HHMM.xlsx
"""
import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

BASE = Path(r"D:\IA MALELO\consultoria\IMPAKTTO")
XLSX = BASE / "CRONOANALISE.xlsx"
SHEET = "4_APONTAMENTO_HIST"
HEADER_ROW = 3
FIRST_DATA_ROW = 4
HEADERS = [
    "DATA","CLIENTE","ID_OP","ITEM","DESCRICAO_OP","N_PROG",
    "QTD_PROD","QTD_PED","H_INICIO","H_FIM","TOTAL_MIN","NC",
    "COD_PARADA","T_PARADA_MIN","OPERADOR","MAQUINA","AUDIT",
]
THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def main(csv_path: Path):
    if not csv_path.exists():
        print(f"ERRO: arquivo nao encontrado: {csv_path}")
        sys.exit(1)

    if not XLSX.exists():
        print(f"ERRO: planilha nao encontrada: {XLSX}")
        sys.exit(1)

    # backup
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    bkp = XLSX.with_name(f"CRONOANALISE_bkp_{stamp}.xlsx")
    shutil.copy2(XLSX, bkp)
    print(f"Backup: {bkp.name}")

    # le csv
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)

    if not rows:
        print("CSV vazio.")
        sys.exit(1)

    head = [c.strip() for c in rows[0]]
    if head != HEADERS:
        print("AVISO: cabecalho diverge do esperado.")
        print(f"  Esperado: {HEADERS}")
        print(f"  Recebido: {head}")
        ans = input("Continuar mesmo assim? [s/N] ").strip().lower()
        if ans != "s":
            sys.exit(1)

    data_rows = rows[1:]
    print(f"Linhas a importar: {len(data_rows)}")

    wb = load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"ERRO: aba {SHEET} nao existe")
        sys.exit(1)
    ws = wb[SHEET]

    # acha primeira linha vazia (coluna A)
    first_empty = FIRST_DATA_ROW
    while ws.cell(row=first_empty, column=1).value not in (None, ""):
        first_empty += 1

    # colunas numericas (1-indexed): C=ID_OP F=N_PROG G=QTD_PROD H=QTD_PED K=TOTAL_MIN L=NC N=T_PARADA_MIN
    NUMERIC_COLS = {3, 6, 7, 8, 11, 12, 14}

    def to_num(v):
        if not v: return None
        s = str(v).strip()
        cj = "cj" in s.lower()
        s2 = s.lower().replace("cj", "").strip().replace(",", ".")
        if not s2 or s2.startswith("?") or "[" in s2: return s
        try:
            f = float(s2)
            n = int(f) if f.is_integer() else f
            return n * 2 if cj else n
        except ValueError:
            return s

    # grava
    for i, row in enumerate(data_rows):
        for j, val in enumerate(row, 1):
            v = val.strip() if val else None
            if v and j in NUMERIC_COLS:
                v = to_num(v)
            cell = ws.cell(row=first_empty + i, column=j, value=v)
            cell.border = BORDER_ALL

    wb.save(XLSX)
    print(f"OK: {len(data_rows)} linhas inseridas a partir da linha {first_empty}.")
    print(f"Arquivo: {XLSX}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_csv.py <arquivo.csv>")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.is_absolute():
        p = BASE / p
    main(p)
