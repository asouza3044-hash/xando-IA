"""
Gerar dados consolidados para o PPTX de Cronoanalise IMPAKTTO.

Le CRONOANALISE.xlsx aba 4_APONTAMENTO_HIST, aplica as limpezas (remover [???],
filtrar operador ilegivel quando necessario para ranking, etc) e salva
dados_relatorio.json para ser consumido por build_pptx.js.

Uso: python gerar_dados.py
"""
from __future__ import annotations
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


AQUI = Path(__file__).resolve().parent
PROJETO = AQUI.parent  # D:\IA MALELO\consultoria\IMPAKTTO
PLANILHA = PROJETO / "CRONOANALISE.xlsx"
SAIDA = AQUI / "dados_relatorio.json"

# Dicionario de correcao: nome no formulario (OCR) -> nome real do operador.
# Em 10/05/2026 a grafia foi corrigida direto na planilha CRONOANALISE.xlsx.
RENOMEAR_OPERADOR = {
    # exemplo: "ulton": "Wendell",
}

# Descricoes dos codigos de parada (rodape do formulario IMPAKTTO).
# Confirmadas com o PDF legivel em 11/05/2026.
PARADA_DESCRICAO = {
    "P01": "Falta de Energia Elétrica",
    "P02": "Falta Material (Matéria Prima)",
    "P03": "Saída de Funcionário",
    "P04": "Troca de Ferramenta",
    "P05": "Manutenção",
    "P06": "Falta Componente",
    "P07": "Almoço",
    "P08": "Preparação",
    "P09": "Inspeção - Controle de Qualidade",
    "P10": "Falta Serviço",
    "P11": "Alarme",
    "P12": "Reunião/Palestra",
    "P13": "Retrabalhos",
    "P14": "Manutenção Preventiva",
    "P15": "Limpeza",
    "P16": "Outras Horas",
}


def limpar_operador(valor):
    """Remove ' [???]' / '[???]' e aplica o dicionario de correcao de nomes."""
    if valor is None:
        return None
    s = str(valor).strip()
    s = re.sub(r"\s*\[\?+\]\s*", "", s)
    s = s.strip()
    if not s:
        return None
    corrigido = RENOMEAR_OPERADOR.get(s.lower())
    return corrigido if corrigido else s


def fmt_data(v):
    """Formata datetime/string/date para 'DD/MM/AAAA'."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def carregar_apontamentos():
    """Le a aba 4_APONTAMENTO_HIST e devolve lista de dicts."""
    if not PLANILHA.exists():
        sys.exit(f"ERRO: planilha nao encontrada em {PLANILHA}")

    wb = load_workbook(PLANILHA, data_only=False)
    ws = wb["4_APONTAMENTO_HIST"]

    # Cabecalho ESTA na linha 3 (linha 1 = titulo, linha 2 = vazia)
    headers = [c.value for c in ws[3]]
    apontamentos = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if all(v is None for v in row):
            continue
        d = dict(zip(headers, row))
        apontamentos.append(d)
    return apontamentos


def num(v):
    """Converte para int/float quando possivel, senao None."""
    if v is None:
        return None
    try:
        if isinstance(v, str):
            v = v.strip().replace(",", ".")
            if v == "" or v == "???":
                return None
        return float(v) if isinstance(v, (int, float, str)) else None
    except (ValueError, TypeError):
        return None


def calcular(apontamentos):
    """Calcula todas as agregacoes que o PPTX precisa."""
    # --- KPIs principais
    total_op = len(apontamentos)
    total_prod = sum(num(a.get("QTD_PROD")) or 0 for a in apontamentos)
    total_ped = sum(num(a.get("QTD_PED")) or 0 for a in apontamentos)
    pct_atendido = (total_prod / total_ped * 100) if total_ped else 0
    paradas_total = sum(1 for a in apontamentos if a.get("COD_PARADA") and str(a["COD_PARADA"]).strip() not in ("", "???"))

    # --- Por operador (consolidando "Nome" e "Nome [???]")
    op_qtd = Counter()
    for a in apontamentos:
        op_limpo = limpar_operador(a.get("OPERADOR"))
        q = num(a.get("QTD_PROD")) or 0
        if op_limpo in (None, "???", "?", ""):
            op_qtd["Operador ilegível"] += q
        else:
            op_qtd[op_limpo] += q
    operadores = [
        {"nome": k, "qtd": int(v), "pct": round(v / total_prod * 100, 1) if total_prod else 0}
        for k, v in op_qtd.most_common()
    ]

    # --- Por peca (ID_OP)
    peca_qtd = Counter()
    for a in apontamentos:
        pid = a.get("ID_OP")
        if pid in (None, ""):
            pid = "(em branco)"
        pid = str(pid).strip()
        q = num(a.get("QTD_PROD")) or 0
        peca_qtd[pid] += q
    pecas = [
        {"id": k, "qtd": int(v), "pct": round(v / total_prod * 100, 1) if total_prod else 0}
        for k, v in peca_qtd.most_common()
    ]

    # --- Pareto de paradas
    par_count = Counter()
    for a in apontamentos:
        cod = a.get("COD_PARADA")
        if cod and str(cod).strip():
            cod_str = str(cod).strip()
            # consolidar "87 [???]" e similares
            cod_limpo = re.sub(r"\s*\[\?+\]\s*", "", cod_str).strip()
            if not cod_limpo:
                cod_limpo = "Cód. ilegível"
            par_count[cod_limpo] += 1
    total_paradas_leg = sum(par_count.values())
    pareto = []
    acum = 0
    for cod, cnt in par_count.most_common():
        acum += cnt
        pareto.append({
            "cod": cod,
            "descricao": PARADA_DESCRICAO.get(cod, "Sem descrição"),
            "ocorr": cnt,
            "pct": round(cnt / total_paradas_leg * 100) if total_paradas_leg else 0,
            "acum": round(acum / total_paradas_leg * 100) if total_paradas_leg else 0,
        })

    # --- Periodo (aceita datetime, date OU string DD/MM/YYYY)
    def parse_data(v):
        if isinstance(v, (datetime, date)):
            return v if isinstance(v, date) else v.date()
        if isinstance(v, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(v.strip(), fmt).date()
                except ValueError:
                    pass
        return None

    datas_parseadas = [d for d in (parse_data(a.get("DATA")) for a in apontamentos) if d is not None]
    if datas_parseadas:
        ini, fim = min(datas_parseadas), max(datas_parseadas)
        periodo = f"{ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
        periodo_curto = f"{ini.strftime('%d')}–{fim.strftime('%d/%m/%Y')}"
    else:
        periodo = "—"
        periodo_curto = "—"

    # --- Clientes
    clientes = Counter()
    for a in apontamentos:
        c = a.get("CLIENTE")
        if c:
            clientes[str(c).strip()] += 1
    cliente_principal = clientes.most_common(1)[0][0] if clientes else "—"

    # --- Atendimento por peca (produzido vs pedido) ordenado por qtd pedida desc
    ped_por_peca = Counter()
    prod_por_peca = Counter()
    ops_por_peca = Counter()
    for a in apontamentos:
        pid = str(a.get("ID_OP") or "—").strip()
        ped_por_peca[pid] += num(a.get("QTD_PED")) or 0
        prod_por_peca[pid] += num(a.get("QTD_PROD")) or 0
        ops_por_peca[pid] += 1
    atendimento_peca = []
    for pid in sorted(ped_por_peca, key=lambda x: -ped_por_peca[x]):
        ped = ped_por_peca[pid]
        prod = prod_por_peca[pid]
        atendimento_peca.append({
            "id": pid,
            "pedido": int(ped),
            "produzido": int(prod),
            "gap": int(ped - prod),
            "pct": round(prod / ped * 100, 1) if ped else 0,
            "ops": ops_por_peca[pid],
        })

    # --- Matriz Operador x Peca (qtd produzida)
    mat = {}
    for a in apontamentos:
        op = limpar_operador(a.get("OPERADOR"))
        if op in (None, "???", ""):
            op = "Ilegível"
        pid = str(a.get("ID_OP") or "—").strip()
        q = num(a.get("QTD_PROD")) or 0
        mat.setdefault(op, {})
        mat[op][pid] = mat[op].get(pid, 0) + int(q)
    matriz_op_peca = {
        "operadores": sorted(mat.keys(),
                              key=lambda o: -sum(mat[o].values())),
        "pecas": sorted({p for d in mat.values() for p in d},
                         key=lambda p: -sum(mat[o].get(p, 0) for o in mat)),
        "celulas": mat,
    }

    # --- Tendencia diaria
    prod_dia = {}
    for a in apontamentos:
        d = parse_data(a.get("DATA"))
        if d is None:
            continue
        q = num(a.get("QTD_PROD")) or 0
        prod_dia[d] = prod_dia.get(d, 0) + int(q)
    tendencia = [
        {"data": d.strftime("%d/%m"), "qtd": v}
        for d, v in sorted(prod_dia.items())
    ]

    # --- Detalhe das OPs (para o slide tabela)
    detalhe = []
    for a in apontamentos:
        operador = limpar_operador(a.get("OPERADOR"))
        if operador in (None, "???", ""):
            operador = "—"
        cod_p = a.get("COD_PARADA")
        cod_p = re.sub(r"\s*\[\?+\]\s*", " [?]", str(cod_p)).strip() if cod_p else "—"
        if cod_p == "" or cod_p == "???":
            cod_p = "—"

        qtd_ped_val = num(a.get("QTD_PED"))
        qtd_prod_val = num(a.get("QTD_PROD"))
        detalhe.append([
            fmt_data(a.get("DATA"))[:5],  # so DD/MM
            str(a.get("CLIENTE") or "—"),
            str(a.get("ID_OP") or "—"),
            str(int(qtd_prod_val)) if qtd_prod_val is not None else "—",
            str(int(qtd_ped_val)) if qtd_ped_val is not None else "—",
            operador,
            cod_p,
        ])

    # Resultado final
    return {
        "geracao": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "kpis": {
            "ops": total_op,
            "produzido": int(total_prod),
            "pedido": int(total_ped),
            "pct_atendido": round(pct_atendido, 1),
            "paradas": paradas_total,
        },
        "periodo": periodo,
        "periodo_curto": periodo_curto,
        "cliente_principal": cliente_principal,
        "operadores": operadores,
        "pecas": pecas,
        "pareto_paradas": pareto,
        "atendimento_peca": atendimento_peca,
        "matriz_op_peca": matriz_op_peca,
        "tendencia_diaria": tendencia,
        "detalhe": detalhe,
    }


def main():
    print(f"Lendo  {PLANILHA}")
    apontamentos = carregar_apontamentos()
    print(f"  {len(apontamentos)} apontamentos carregados")
    dados = calcular(apontamentos)

    SAIDA.write_text(json.dumps(dados, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"OK  dados consolidados salvos em {SAIDA}")
    print(f"  KPIs: {dados['kpis']['ops']} OPs, "
          f"{dados['kpis']['produzido']} pcs, "
          f"{dados['kpis']['pct_atendido']}% atendido, "
          f"{dados['kpis']['paradas']} paradas")
    print(f"  Periodo: {dados['periodo']}")
    print(f"  Cliente principal: {dados['cliente_principal']}")
    if dados['operadores']:
        top = dados['operadores'][0]
        print(f"  Top operador: {top['nome']} ({top['qtd']} pcs)")


if __name__ == "__main__":
    main()
