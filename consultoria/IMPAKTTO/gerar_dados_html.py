"""
gerar_dados_html.py
Le CRONOANALISE.xlsx e injeta os dados no dashboard HTML v2.
Execute sempre que importar novos apontamentos:
    python gerar_dados_html.py
"""

import openpyxl
import json
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

BASE     = Path(__file__).parent
EXCEL    = BASE / "CRONOANALISE.xlsx"
OUTDIR   = BASE / "entregaveis"
TEMPLATE = OUTDIR / "CRONOANALISE_IMPAKTTO_v2.html"
OUTDIR.mkdir(exist_ok=True)

# --- LER EXCEL ---
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb["4_APONTAMENTO_HIST"]
headers = [ws.cell(3, c).value for c in range(1, 18)]

rows = []
for r in range(4, 2000):
    if ws.cell(r, 1).value is None:
        break
    row = {headers[c-1]: ws.cell(r, c).value for c in range(1, 18)}
    rows.append(row)

print("Linhas lidas: " + str(len(rows)))

# --- FUNCOES LIMPEZA ---
_CLIENTES_MAP = {"matro": "MATÃO", "matao": "MATÃO", "matão": "MATÃO"}

_MAQUINAS_MAP = {
    "cepha usinagem": "Centro Usinagem",
    "cepha":          "Centro Usinagem",
    "cepho":          "Centro Usinagem",
    "centro":         "Centro Usinagem",
}

_OPERADORES_MAP = {
    "cleiton": "Clayton",
}

def limpar_str(v):
    if v is None: return ""
    s = str(v).strip().replace("[???]","").replace("???","").replace("[?]","").strip()
    lower = s.lower()
    if lower in _CLIENTES_MAP:   return _CLIENTES_MAP[lower]
    if lower in _MAQUINAS_MAP:   return _MAQUINAS_MAP[lower]
    if lower in _OPERADORES_MAP: return _OPERADORES_MAP[lower]
    return s

def limpar_int(v):
    if v is None: return 0
    s = str(v).strip().lower()
    cj = "cj" in s
    s = s.replace("cj", "").strip()
    try:
        n = int(float(s.replace(",", ".")))
        return n * 2 if cj else n
    except: return 0

def limpar_cod(v):
    s = limpar_str(v)
    return s if re.match(r'^P\d+$', s) else None

def parse_hora(v):
    if v is None: return None
    s = str(v).strip()
    parts = re.split(r'[:\.]', s)
    if len(parts) >= 2:
        try:
            return int(parts[0]) + int(parts[1]) / 60
        except: pass
    return None

def pch_calc(h_i, h_f, qtd):
    if h_i is None or h_f is None or qtd <= 0: return None
    dur = h_f - h_i
    if dur <= 0 or dur > 24: return None
    return round(qtd / dur, 1)

# --- AGRUPAMENTOS ---
total_prod = total_ped = paradas_total = 0
datas = []
clientes_set = set()
maquinas_set = set()

qtd_op     = defaultdict(int)
ops_op     = defaultdict(int)
pch_op_sum = defaultdict(float)
pch_op_n   = defaultdict(int)

qtd_peca   = defaultdict(int)
ped_peca   = defaultdict(int)
qtd_parada = defaultdict(int)
qtd_dia    = defaultdict(int)
qtd_maq    = defaultdict(int)
ops_maq    = defaultdict(int)

for r in rows:
    op     = limpar_str(r.get("OPERADOR"))
    peca   = limpar_str(r.get("ID_OP"))
    parada = limpar_cod(r.get("COD_PARADA"))
    qp     = limpar_int(r.get("QTD_PROD"))
    qd     = limpar_int(r.get("QTD_PED"))
    data   = limpar_str(r.get("DATA"))
    cli    = limpar_str(r.get("CLIENTE"))
    maq    = limpar_str(r.get("MAQUINA"))
    h_i    = parse_hora(r.get("H_INICIO"))
    h_f    = parse_hora(r.get("H_FIM"))
    pch    = pch_calc(h_i, h_f, qp)

    total_prod += qp
    total_ped  += qd
    if cli: clientes_set.add(cli)
    if maq: maquinas_set.add(maq)

    if op:
        qtd_op[op]  += qp
        ops_op[op]  += 1
        if pch:
            pch_op_sum[op] += pch
            pch_op_n[op]   += 1

    if peca and peca.lower() != "teste":
        qtd_peca[peca] += qp
        ped_peca[peca] += qd

    if parada:
        qtd_parada[parada] += 1
        paradas_total += 1

    if maq:
        qtd_maq[maq] += qp
        ops_maq[maq] += 1

    if data and len(data) >= 8:
        p = data.split("/")
        if len(p) == 3:
            sort_k = p[2] + "-" + p[1] + "-" + p[0]
            show_k = p[0] + "/" + p[1]
            qtd_dia[(sort_k, show_k)] += qp
            datas.append(data)

# --- MONTAR DADOS ---
pct_atendido = round(total_prod / total_ped * 100, 1) if total_ped else 0

operadores_list = sorted([
    {
        "nome": k,
        "qtd":  v,
        "ops":  ops_op[k],
        "pct":  round(v / total_prod * 100, 1) if total_prod else 0,
        "pch":  round(pch_op_sum[k] / pch_op_n[k], 1) if pch_op_n[k] else None,
    }
    for k, v in qtd_op.items() if k
], key=lambda x: -x["qtd"])

pecas_list = sorted([
    {"id": k, "qtd": v, "ped": ped_peca[k],
     "pct": round(v / total_prod * 100, 1) if total_prod else 0}
    for k, v in qtd_peca.items() if k
], key=lambda x: -x["qtd"])

parada_sort = sorted(qtd_parada.items(), key=lambda x: -x[1])
acum = 0
pareto_list = []
for cod, ocorr in parada_sort:
    acum += ocorr
    pareto_list.append({"cod": cod, "ocorr": ocorr,
                        "pct_acum": round(acum / paradas_total * 100, 1) if paradas_total else 0})

tendencia_list = [
    {"dia": show, "qtd": qtd}
    for (sk, show), qtd in sorted(qtd_dia.items(), key=lambda x: x[0][0])
]

maquinas_list = sorted([
    {"nome": k, "qtd": v, "ops": ops_maq[k]}
    for k, v in qtd_maq.items() if k
], key=lambda x: -x["qtd"])

def data_iso(s):
    p = s.split("/")
    return f"{p[2]}-{p[1]}-{p[0]}" if len(p) == 3 else s

periodo = (min(datas, key=data_iso) + " a " + max(datas, key=data_iso)) if datas else ""

ops_lista = []
for r in rows:
    ops_lista.append({
        "data":     limpar_str(r.get("DATA")),
        "cliente":  limpar_str(r.get("CLIENTE")),
        "id_op":    limpar_str(r.get("ID_OP")),
        "item":     limpar_str(r.get("ITEM")),
        "desc":     limpar_str(r.get("DESCRICAO_OP")),
        "n_prog":   limpar_str(r.get("N_PROG")),
        "qtd_prod": limpar_int(r.get("QTD_PROD")),
        "qtd_ped":  limpar_int(r.get("QTD_PED")),
        "h_inicio": limpar_str(r.get("H_INICIO")),
        "h_fim":    limpar_str(r.get("H_FIM")),
        "operador": limpar_str(r.get("OPERADOR")),
        "maquina":  limpar_str(r.get("MAQUINA")),
        "parada":   limpar_str(r.get("COD_PARADA")),
        "audit":    limpar_str(r.get("AUDIT")),
    })

dados = {
    "ops":          len(rows),
    "total_prod":   total_prod,
    "total_ped":    total_ped,
    "pct_atendido": pct_atendido,
    "paradas":      paradas_total,
    "clientes":     sorted(list(clientes_set)),
    "maquinas":     sorted(list(maquinas_set)),
    "periodo":      periodo,
    "operadores":   operadores_list,
    "pecas":        pecas_list,
    "parada_lista": pareto_list,
    "tendencia":    tendencia_list,
    "maquinas_list":maquinas_list,
    "ops_lista":    ops_lista,
}

dados_json = json.dumps(dados, ensure_ascii=False, indent=2)
agora = datetime.now().strftime("%d/%m/%Y %H:%M")

# --- INJETAR NO HTML ---
if not TEMPLATE.exists():
    print("AVISO: Template v2 nao encontrado: " + str(TEMPLATE))
    print("Crie o arquivo entregaveis/CRONOANALISE_IMPAKTTO_v2.html primeiro.")
    exit(1)

html = TEMPLATE.read_text(encoding="utf-8")
MARK_START = "// @@DADOS_START@@"
MARK_END   = "// @@DADOS_END@@"

if MARK_START not in html or MARK_END not in html:
    print("Marcadores nao encontrados no template.")
    exit(1)

bloco = (MARK_START + "\n"
         + "// AUTO-GERADO em " + agora + " — " + str(len(rows)) + " registros\n"
         + "const DADOS = " + dados_json + ";\n"
         + MARK_END)

antes  = html.split(MARK_START)[0]
depois = html.split(MARK_END)[1]
html   = antes + bloco + depois

TEMPLATE.write_text(html, encoding="utf-8")

print("Dashboard v2 atualizado: " + str(TEMPLATE.name))
print(f"OPs:        {len(rows)}")
print(f"Produzido:  {total_prod} peças")
print(f"Atendido:   {pct_atendido}%")
print(f"Paradas:    {paradas_total}")
print(f"Operadores: {len(operadores_list)}")
print(f"Máquinas:   {len(maquinas_set)}")
print(f"Período:    {periodo}")
print()
print("Abra no navegador:")
print(str(TEMPLATE))
