"""
Gera CRONOANALISE.xlsx — versao 2 (visual de consultoria).
Capa, dashboard com graficos, KPI cards, paleta corporativa.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.fill import ColorChoice
from pathlib import Path

OUT = Path(r"D:\IA MALELO\consultoria\IMPAKTTO\CRONOANALISE.xlsx")

# ==================== PALETA CORPORATIVA ====================
NAVY    = "1F3A5F"   # azul-marinho — titulos, headers
GRAPH   = "3C4858"   # grafite — texto
ORANGE  = "E87722"   # accent / alerta
GREEN   = "2E933C"   # bom / meta
RED     = "C0392B"   # ruim / abaixo meta
LIGHT   = "F5F7FA"   # fundo claro
HEADER2 = "2E5984"   # azul medio (sub-headers)
GREY    = "BFBFBF"   # bordas
WHITE   = "FFFFFF"

FONT_NAME = "Calibri"

# ==================== ESTILOS ====================
def f(size=11, bold=False, color=GRAPH, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

THIN = Side(border_style="thin", color=GREY)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=1)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

def style_header_row(ws, row, ncols, color=NAVY):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(color)
        cell.font = f(11, bold=True, color=WHITE)
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 28

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def borders_block(ws, r1, c1, r2, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = BORDER_ALL

def hide_grid(ws):
    ws.sheet_view.showGridLines = False

# =====================================================================
wb = Workbook()
wb.remove(wb.active)

# =====================================================================
# 0_CAPA — banda lateral azul-marinho, titulo grande, info projeto
# =====================================================================
ws = wb.create_sheet("0_CAPA")
hide_grid(ws)
set_widths(ws, [3, 2, 30, 30, 30, 30, 30, 3])
for r in range(1, 41):
    ws.row_dimensions[r].height = 22

# banda lateral azul-marinho (col A)
for r in range(1, 41):
    ws.cell(row=r, column=1).fill = fill(NAVY)

# titulo
ws["C4"] = "CRONOANÁLISE INDUSTRIAL"
ws["C4"].font = f(28, bold=True, color=NAVY)
ws.merge_cells("C4:G4")

ws["C5"] = "Diagnóstico de capacidade, produtividade e tempo-padrão"
ws["C5"].font = f(13, italic=True, color=GRAPH)
ws.merge_cells("C5:G5")

# linha laranja
for c in range(3, 8):
    ws.cell(row=7, column=c).fill = fill(ORANGE)
ws.row_dimensions[7].height = 4

# bloco identificacao
labels = [
    ("CLIENTE",   "IMPAKTTO"),
    ("CONSULTOR", "Alexandre Souza — LASEC"),
    ("CONTATO",   "(11) 3936-5041 · orcamento@lasec.com.br"),
    ("INÍCIO",    "2026-05-09"),
    ("ESCOPO",    "Fábrica completa — usinagem CNC"),
    ("MÉTODO",    "Apontamento histórico (OCR) + Cronometragem + Amostragem do Trabalho"),
    ("ENTREGÁVEIS", "Tempo-Padrão · Carga-Máquina · Produtividade O×M×P · Plano de Melhoria"),
]
r = 10
for label, value in labels:
    cl = ws.cell(row=r, column=3, value=label)
    cl.font = f(10, bold=True, color=ORANGE)
    cl.alignment = LEFT
    cv = ws.cell(row=r, column=4, value=value)
    cv.font = f(12, color=GRAPH)
    cv.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    r += 1

# bloco SUMARIO
ws["C20"] = "SUMÁRIO"
ws["C20"].font = f(14, bold=True, color=NAVY)

abas = [
    ("1",  "DASHBOARD",        "Visão executiva — KPIs, Pareto de paradas, ocupação"),
    ("2",  "INSTRUÇÕES",       "Guia de uso e fluxo metodológico"),
    ("3",  "CADASTROS",        "Máquinas, operadores, peças, códigos parada, turnos"),
    ("4",  "APONTAMENTO_HIST", "Histórico do formulário IMPAKTTO (via OCR)"),
    ("5",  "COLETA_CRONO",     "Cronometragem de calibração (10 ciclos × peça)"),
    ("6",  "COLETA_AMOSTRAGEM","Amostragem do trabalho (work sampling)"),
    ("7",  "TEMPOS_PADRAO",    "TP oficial — Fator de Apontamento × peças/hora"),
    ("8",  "CARGA_MAQUINA",    "Capacidade × demanda × ocupação real"),
    ("9",  "PRODUTIVIDADE",    "Matriz Operador × Máquina × Peça"),
    ("10", "PLANO_MELHORIA",   "Ações priorizadas (impacto × esforço)"),
]
r = 22
for idx, nome, desc in abas:
    ci = ws.cell(row=r, column=3, value=idx)
    ci.font = f(11, bold=True, color=ORANGE)
    ci.alignment = CENTER
    cn = ws.cell(row=r, column=4, value=nome)
    cn.font = f(11, bold=True, color=NAVY)
    cn.alignment = LEFT
    cd = ws.cell(row=r, column=5, value=desc)
    cd.font = f(10, color=GRAPH)
    cd.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    # linha de fundo cinza claro alternada
    if r % 2 == 0:
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
    r += 1

# rodape
ws["C38"] = f"Versão 2.0 · 2026-05-09 · LASEC Consultoria"
ws["C38"].font = f(9, italic=True, color=GRAPH)
ws.merge_cells("C38:G38")

# =====================================================================
# 1_DASHBOARD — KPI cards + graficos
# =====================================================================
ws = wb.create_sheet("1_DASHBOARD")
hide_grid(ws)
set_widths(ws, [2, 18, 18, 18, 18, 18, 18, 18, 18, 2])
for r in range(1, 51):
    ws.row_dimensions[r].height = 20

# cabecalho dashboard
ws["B2"] = "DASHBOARD EXECUTIVO"
ws["B2"].font = f(20, bold=True, color=NAVY)
ws.merge_cells("B2:I2")

ws["B3"] = "IMPAKTTO · Cronoanálise · atualização automática conforme dados são preenchidos"
ws["B3"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("B3:I3")

# linha laranja
for c in range(2, 10):
    ws.cell(row=4, column=c).fill = fill(ORANGE)
ws.row_dimensions[4].height = 3

# === 4 KPI CARDS ===
# Card 1: OEE estimado | Card 2: Ocupacao media | Card 3: % Paradas evitaveis | Card 4: FA medio
def kpi_card(ws, col, label, value_formula, sub, color, fmt="0"):
    """col = coluna inicial (2 colunas largas)."""
    r1 = 6
    # banda superior colorida
    for c in (col, col+1):
        ws.cell(row=r1, column=c).fill = fill(color)
    ws.row_dimensions[r1].height = 6
    # label
    cl = ws.cell(row=r1+1, column=col, value=label)
    cl.font = f(9, bold=True, color=GRAPH)
    cl.alignment = CENTER
    cl.fill = fill(LIGHT)
    ws.merge_cells(start_row=r1+1, start_column=col, end_row=r1+1, end_column=col+1)
    # valor grande
    cv = ws.cell(row=r1+2, column=col, value=value_formula)
    cv.font = f(28, bold=True, color=NAVY)
    cv.alignment = CENTER
    cv.fill = fill(LIGHT)
    cv.number_format = fmt
    ws.merge_cells(start_row=r1+2, start_column=col, end_row=r1+3, end_column=col+1)
    ws.row_dimensions[r1+2].height = 30
    ws.row_dimensions[r1+3].height = 18
    # subtitle
    cs = ws.cell(row=r1+4, column=col, value=sub)
    cs.font = f(9, italic=True, color=GRAPH)
    cs.alignment = CENTER
    cs.fill = fill(LIGHT)
    ws.merge_cells(start_row=r1+4, start_column=col, end_row=r1+4, end_column=col+1)
    # bordas
    borders_block(ws, r1, col, r1+4, col+1)

kpi_card(ws, 2, "OPs REGISTRADAS",       "=IFERROR(_AGREG!B2,\"—\")", "no apontamento OCR",     NAVY,   fmt="0")
kpi_card(ws, 4, "QTD PRODUZIDA",         "=IFERROR(_AGREG!B3,\"—\")", "soma das peças",         GREEN,  fmt="#,##0")
kpi_card(ws, 6, "% ATENDIDO",            "=IFERROR(_AGREG!B4,\"—\")", "produzido ÷ pedido",     ORANGE, fmt="0.0%")
kpi_card(ws, 8, "PARADAS REGISTRADAS",   "=IFERROR(_AGREG!B5,\"—\")", "ocorrências códigos P", RED,    fmt="0")

# =====================================================
# SECAO 1: PRODUTIVIDADE POR OPERADOR  (FOCO PRINCIPAL)
# =====================================================
ws["B14"] = "PRODUTIVIDADE POR OPERADOR"
ws["B14"].font = f(14, bold=True, color=NAVY)
ws.merge_cells("B14:E14")
ws["B15"] = "OPs · Quantidade produzida · % do mix de produção"
ws["B15"].font = f(9, italic=True, color=GRAPH)
ws.merge_cells("B15:E15")

op_headers = ["OPERADOR", "OPs", "QTD_PROD", "% MIX"]
for i, h in enumerate(op_headers, 2):
    ws.cell(row=17, column=i, value=h)
style_header_row(ws, 17, len(op_headers))
for r in range(18, 23):  # 5 operadores
    ws.cell(row=r, column=2, value=f"=IFERROR(_AGREG!D{r-15},\"\")")
    ws.cell(row=r, column=3, value=f"=IFERROR(_AGREG!E{r-15},\"\")")
    ws.cell(row=r, column=4, value=f"=IFERROR(_AGREG!F{r-15},\"\")")
    ws.cell(row=r, column=5, value=f"=IFERROR(_AGREG!G{r-15},\"\")")
    ws.cell(row=r, column=5).number_format = "0.0%"
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = BORDER_ALL
    if r % 2 == 0:
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
ws.conditional_formatting.add(
    "E18:E22",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=NAVY)
)

# =====================================================
# SECAO 2: PRODUTIVIDADE POR PECA  (FOCO PRINCIPAL)
# =====================================================
ws["G14"] = "PRODUTIVIDADE POR PEÇA"
ws["G14"].font = f(14, bold=True, color=NAVY)
ws.merge_cells("G14:I14")
ws["G15"] = "Mix produzido — base item do formulário"
ws["G15"].font = f(9, italic=True, color=GRAPH)
ws.merge_cells("G15:I15")

pe_headers = ["ITEM", "QTD_PROD", "% MIX"]
for i, h in enumerate(pe_headers, 7):
    ws.cell(row=17, column=i, value=h)
for c in range(7, 10):
    cell = ws.cell(row=17, column=c)
    cell.fill = fill(NAVY)
    cell.font = f(11, bold=True, color=WHITE)
    cell.alignment = CENTER
    cell.border = BORDER_ALL
ws.row_dimensions[17].height = 28

for r in range(18, 22):  # 4 pecas
    ws.cell(row=r, column=7, value=f"=IFERROR(_AGREG!K{r-15},\"\")")
    ws.cell(row=r, column=8, value=f"=IFERROR(_AGREG!M{r-15},\"\")")
    ws.cell(row=r, column=9, value=f"=IFERROR(_AGREG!N{r-15},\"\")")
    ws.cell(row=r, column=9).number_format = "0.0%"
    for c in range(7, 10):
        ws.cell(row=r, column=c).border = BORDER_ALL
    if r % 2 == 0:
        for c in range(7, 10):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
ws.conditional_formatting.add(
    "I18:I21",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ORANGE)
)

# === GRAFICO 1: barras horizontais — produtividade por operador ===
chart_op = BarChart()
chart_op.type = "bar"
chart_op.style = 12
chart_op.title = "Quantidade produzida por operador"
data = Reference(ws, min_col=4, min_row=17, max_col=4, max_row=22)
cats = Reference(ws, min_col=2, min_row=18, max_col=2, max_row=22)
chart_op.add_data(data, titles_from_data=True)
chart_op.set_categories(cats)
chart_op.dLbls = DataLabelList(showVal=True)
chart_op.height = 9
chart_op.width = 16
ws.add_chart(chart_op, "B25")

# === GRAFICO 2: barras — produtividade por peça ===
chart_pe = BarChart()
chart_pe.type = "col"
chart_pe.style = 13
chart_pe.title = "Quantidade produzida por peça"
data = Reference(ws, min_col=8, min_row=17, max_col=8, max_row=21)
cats = Reference(ws, min_col=7, min_row=18, max_col=7, max_row=21)
chart_pe.add_data(data, titles_from_data=True)
chart_pe.set_categories(cats)
chart_pe.dLbls = DataLabelList(showVal=True)
chart_pe.height = 9
chart_pe.width = 16
ws.add_chart(chart_pe, "G25")

# =====================================================
# SECAO 3: PARETO PARADAS  (secundário, abaixo)
# =====================================================
ws["B45"] = "PARETO DE PARADAS (frequência)"
ws["B45"].font = f(14, bold=True, color=NAVY)
ws.merge_cells("B45:E45")
ws["B46"] = "Códigos vistos no formulário — frequência (não horas, ainda)"
ws["B46"].font = f(9, italic=True, color=GRAPH)
ws.merge_cells("B46:E46")

par_headers = ["CÓDIGO", "OCORRÊNCIAS", "% ACUM"]
for i, h in enumerate(par_headers, 2):
    ws.cell(row=48, column=i, value=h)
style_header_row(ws, 48, len(par_headers))

for r in range(49, 53):  # 4 codigos
    ws.cell(row=r, column=2, value=f"=IFERROR(_AGREG!P{r-46},\"\")")
    ws.cell(row=r, column=3, value=f"=IFERROR(_AGREG!Q{r-46},\"\")")
    ws.cell(row=r, column=4, value=f"=IFERROR(_AGREG!R{r-46},\"\")")
    ws.cell(row=r, column=4).number_format = "0.0%"
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER_ALL
    if r % 2 == 0:
        for c in range(2, 5):
            ws.cell(row=r, column=c).fill = fill(LIGHT)

chart_par = BarChart()
chart_par.type = "col"
chart_par.style = 11
chart_par.title = "Frequência de paradas"
data = Reference(ws, min_col=3, min_row=48, max_col=3, max_row=52)
cats = Reference(ws, min_col=2, min_row=49, max_col=2, max_row=52)
chart_par.add_data(data, titles_from_data=True)
chart_par.set_categories(cats)
chart_par.dLbls = DataLabelList(showVal=True)
chart_par.height = 8
chart_par.width = 14
ws.add_chart(chart_par, "F45")

# =====================================================================
# _AGREG — agregacoes DINAMICAS dos dados reais (4_APONTAMENTO_HIST)
# Cols: A=DATA B=CLIENTE C=ID_OP D=ITEM E=DESC F=N_PROG G=QTD_PROD H=QTD_PED
#       I=H_INI J=H_FIM K=TOT_MIN L=NC M=COD_PARADA N=T_PARADA O=OPERADOR P=MAQ Q=AUDIT
# =====================================================================
ws = wb.create_sheet("_AGREG")
hide_grid(ws)
APO = "'4_APONTAMENTO_HIST'"
RNG_ITEM = f"{APO}!$C$4:$C$203"
RNG_PROG = f"{APO}!$F$4:$F$203"
RNG_QP   = f"{APO}!$G$4:$G$203"
RNG_QPD  = f"{APO}!$H$4:$H$203"
RNG_PAR  = f"{APO}!$M$4:$M$203"
RNG_OP   = f"{APO}!$O$4:$O$203"

# === KPIs ===
ws["A1"] = "KPIs (dinâmico)"
ws["A1"].font = f(11, bold=True, color=NAVY)
kpis = [
    ("OPs registradas",       f'=COUNTA({APO}!$A$4:$A$203)'),
    ("Total qtd produzida",   f'=SUM({RNG_QP})'),
    ("% atendido (prod/ped)", f'=IFERROR(SUM({RNG_QP})/SUM({RNG_QPD}),0)'),
    ("Paradas registradas",   f'=SUMPRODUCT(--({RNG_PAR}<>""))'),
]
for i, (lbl, frm) in enumerate(kpis, 2):
    ws.cell(row=i, column=1, value=lbl)
    ws.cell(row=i, column=2, value=frm)
ws["B4"].number_format = "0.0%"

# === PRODUCAO POR OPERADOR (cols D..G) ===
ws["D1"] = "PRODUTIVIDADE — POR OPERADOR"
ws["D1"].font = f(11, bold=True, color=NAVY)
ws.cell(row=2, column=4, value="OPERADOR")
ws.cell(row=2, column=5, value="OPs")
ws.cell(row=2, column=6, value="QTD_PROD")
ws.cell(row=2, column=7, value="% MIX")
operadores = ["Clayton","Ulton","Vanildo","Vita","dilma"]
ws["I2"] = f'=SUM({RNG_QP})'  # total auxiliar
for i, op in enumerate(operadores, 3):
    ws.cell(row=i, column=4, value=op)
    ws.cell(row=i, column=5, value=f'=COUNTIF({RNG_OP},"*{op}*")')
    ws.cell(row=i, column=6, value=f'=SUMIF({RNG_OP},"*{op}*",{RNG_QP})')
    ws.cell(row=i, column=7,
            value=f'=IFERROR(F{i}/MAX($I$2,1),0)')
    ws.cell(row=i, column=7).number_format = "0.0%"

# === PRODUCAO POR PECA / ITEM (cols K..N) ===
ws["K1"] = "PRODUTIVIDADE — POR PEÇA"
ws["K1"].font = f(11, bold=True, color=NAVY)
ws.cell(row=2, column=11, value="ITEM")
ws.cell(row=2, column=12, value="OPs")
ws.cell(row=2, column=13, value="QTD_PROD")
ws.cell(row=2, column=14, value="% MIX")
itens_reais = [30809, 31552, 31652, 31553]
for i, it in enumerate(itens_reais, 3):
    ws.cell(row=i, column=11, value=it)
    ws.cell(row=i, column=12, value=f'=COUNTIF({RNG_ITEM},{it})')
    ws.cell(row=i, column=13, value=f'=SUMIF({RNG_ITEM},{it},{RNG_QP})')
    ws.cell(row=i, column=14,
            value=f'=IFERROR(M{i}/MAX($I$2,1),0)')
    ws.cell(row=i, column=14).number_format = "0.0%"

# === PARETO PARADAS (cols P..S) — secundario ===
ws["P1"] = "Pareto Paradas (frequência)"
ws["P1"].font = f(11, bold=True, color=NAVY)
ws.cell(row=2, column=16, value="CÓDIGO")
ws.cell(row=2, column=17, value="OCORR")
ws.cell(row=2, column=18, value="% ACUM")
codigos_reais = ["P01","P07","P11","P16"]
for i, cod in enumerate(codigos_reais, 3):
    ws.cell(row=i, column=16, value=cod)
    ws.cell(row=i, column=17, value=f'=COUNTIF({RNG_PAR},"{cod}")')
    ws.cell(row=i, column=18,
            value=f'=IFERROR(SUM($Q$3:Q{i})/SUM($Q$3:$Q$10),0)')
    ws.cell(row=i, column=18).number_format = "0.0%"

ws["A15"] = "ℹ Tudo calculado de 4_APONTAMENTO_HIST. Atualiza com novas fotos."
ws["A15"].font = f(9, italic=True, color=ORANGE)

ws.sheet_state = "hidden"

# =====================================================================
# 2_INSTRUCOES
# =====================================================================
ws = wb.create_sheet("2_INSTRUCOES")
hide_grid(ws)
set_widths(ws, [3, 22, 70])

ws["B2"] = "INSTRUÇÕES E FLUXO"
ws["B2"].font = f(20, bold=True, color=NAVY)
ws.merge_cells("B2:C2")
for c in (2, 3):
    ws.cell(row=4, column=c).fill = fill(ORANGE)
ws.row_dimensions[4].height = 3

blocos = [
    ("OBJETIVO",   "Cronoanálise IMPAKTTO → Tempo-Padrão, Carga-Máquina, Produtividade O×M×P, Plano de Melhoria."),
    ("FONTE PRIMÁRIA", "Apontamento manual existente (formulário IMPAKTTO) → digitalizado via OCR Claude Vision na aba 4_APONTAMENTO_HIST."),
    ("FONTES SECUNDÁRIAS", "5_COLETA_CRONO (cronometragem de calibração) e 6_COLETA_AMOSTRAGEM (work sampling)."),
]
r = 6
for tag, txt in blocos:
    cl = ws.cell(row=r, column=2, value=tag)
    cl.font = f(10, bold=True, color=ORANGE); cl.alignment = LEFT
    cv = ws.cell(row=r, column=3, value=txt)
    cv.font = f(11, color=GRAPH); cv.alignment = LEFT
    ws.row_dimensions[r].height = 28
    r += 1

ws.cell(row=r+1, column=2, value="FLUXO METODOLÓGICO").font = f(13, bold=True, color=NAVY)
r += 3
fluxo = [
    "1", "Preencher 3_CADASTROS (máquinas, operadores, peças, códigos parada, turnos).",
    "2", "Importar histórico do formulário IMPAKTTO via OCR → 4_APONTAMENTO_HIST (60 dias).",
    "3", "Cronometrar peças Pareto (5–8 peças, 10 ciclos cada) → 5_COLETA_CRONO.",
    "4", "Amostragem do trabalho (~400 obs/máquina) → 6_COLETA_AMOSTRAGEM.",
    "5", "TP gerado em 7_TEMPOS_PADRAO consolida 5 + valida com 4.",
    "6", "8_CARGA_MAQUINA: capacidade × demanda × ocupação real.",
    "7", "9_PRODUTIVIDADE: matriz Operador × Máquina × Peça + Fator de Apontamento.",
    "8", "1_DASHBOARD atualiza automaticamente; revisar e validar com cliente.",
    "9", "10_PLANO_MELHORIA: priorizar ações por impacto/esforço.",
]
for i in range(0, len(fluxo), 2):
    cl = ws.cell(row=r, column=2, value=fluxo[i])
    cl.font = f(11, bold=True, color=ORANGE); cl.alignment = CENTER
    cv = ws.cell(row=r, column=3, value=fluxo[i+1])
    cv.font = f(11, color=GRAPH); cv.alignment = LEFT
    if r % 2 == 0:
        ws.cell(row=r, column=2).fill = fill(LIGHT)
        ws.cell(row=r, column=3).fill = fill(LIGHT)
    r += 1

r += 2
ws.cell(row=r, column=2, value="REGRAS").font = f(13, bold=True, color=NAVY)
r += 1
regras = [
    "Tempos sempre em MINUTOS decimais (7,5 = 7 min 30 s).",
    "Datas DD/MM/AAAA.",
    "Códigos de parada padronizados pela aba 3 (lista P01..Pxx).",
    "Coluna AUDIT em 4 obrigatória antes de entrar em análise (OK / AJ / ??).",
    "Auditoria amostral 10% das linhas digitalizadas.",
    "Nunca apagar cadastro com vínculo ativo — usar coluna ATIVO=N.",
]
for txt in regras:
    cl = ws.cell(row=r, column=2, value="•")
    cl.font = f(11, bold=True, color=ORANGE); cl.alignment = CENTER
    cv = ws.cell(row=r, column=3, value=txt)
    cv.font = f(11, color=GRAPH); cv.alignment = LEFT
    r += 1

# =====================================================================
# 3_CADASTROS
# =====================================================================
ws = wb.create_sheet("3_CADASTROS")
hide_grid(ws)
set_widths(ws, [16, 28, 16, 16, 16, 16, 16, 28])

ws["A1"] = "CADASTROS"
ws["A1"].font = f(20, bold=True, color=NAVY)
ws.merge_cells("A1:H1")
for c in range(1, 9):
    ws.cell(row=2, column=c).fill = fill(ORANGE)
ws.row_dimensions[2].height = 3

def section(ws, row, title, headers, nrows, color=NAVY):
    ws.cell(row=row, column=1, value=title).font = f(13, bold=True, color=color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(row=row+1, column=i, value=h)
    style_header_row(ws, row+1, len(headers), color=color)
    borders_block(ws, row+2, 1, row+1+nrows, len(headers))
    # zebra
    for r in range(row+2, row+2+nrows):
        if (r - (row+2)) % 2 == 1:
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = fill(LIGHT)

section(ws, 4,  "MÁQUINAS",
        ["COD_MAQ","NOME","TIPO","CÉLULA","EIXOS","HORAS_DIA","TURNOS","ATIVO"], 12)
section(ws, 20, "OPERADORES",
        ["COD_OP","NOME","FUNÇÃO","TURNO","ADMISSÃO","MAQ_PRIM","ATIVO","OBS"], 18)
section(ws, 42, "PEÇAS",
        ["COD_PECA","DESCRIÇÃO","CLIENTE","MATERIAL","PESO_KG","MAQ_PRINC","N_PROG","ATIVO"], 25)
section(ws, 71, "CÓDIGOS DE PARADA",
        ["CÓDIGO","DESCRIÇÃO","CATEGORIA","EVITÁVEL","RESPONSÁVEL","OBS","",""], 15)
# pre-preenche P codes
# SOMENTE codigos vistos na foto real IMPAKTTO (06-15/04/2026)
# Descricao a confirmar com cliente — preenchimento provisorio
codes_seed = [
    ("P01", "(confirmar com IMPAKTTO)", "?", "?", "?"),
    ("P07", "(confirmar com IMPAKTTO)", "?", "?", "?"),
    ("P11", "(confirmar com IMPAKTTO)", "?", "?", "?"),
    ("P16", "(confirmar com IMPAKTTO)", "?", "?", "?"),
]
for i, row in enumerate(codes_seed, 73):
    for j, v in enumerate(row, 1):
        ws.cell(row=i, column=j, value=v)

section(ws, 90, "TURNOS",
        ["TURNO","INÍCIO","FIM","INTERVALO_MIN","DIAS","HORAS_LIQ","",""], 4)
# Turnos a confirmar com IMPAKTTO (foto mostra OPs 07h-19h e 20h-06h, sugestivo 2 turnos)
turnos_seed = []
for i, row in enumerate(turnos_seed, 92):
    for j, v in enumerate(row, 1):
        ws.cell(row=i, column=j, value=v)

# =====================================================================
# 4_APONTAMENTO_HIST
# =====================================================================
ws = wb.create_sheet("4_APONTAMENTO_HIST")
hide_grid(ws)
set_widths(ws, [11, 14, 12, 12, 30, 10, 10, 10, 9, 9, 11, 8, 11, 12, 14, 14, 8])

ws["A1"] = "APONTAMENTO HISTÓRICO (OCR formulário IMPAKTTO)"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:Q1")

headers = [
    "DATA","CLIENTE","ID_OP","ITEM","DESCRICAO_OP","N_PROG",
    "QTD_PROD","QTD_PED","H_INICIO","H_FIM","TOTAL_MIN","NC",
    "COD_PARADA","T_PARADA_MIN","OPERADOR","MAQUINA","AUDIT"
]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header_row(ws, 3, len(headers))
borders_block(ws, 4, 1, 203, len(headers))
# zebra
for r in range(4, 204):
    if (r-4) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
# validacao AUDIT
dv = DataValidation(type="list", formula1='"OK,AJ,??"', allow_blank=True)
dv.add(f"Q4:Q203")
ws.add_data_validation(dv)
ws.freeze_panes = "A4"

# =====================================================================
# 5_COLETA_CRONO
# =====================================================================
ws = wb.create_sheet("5_COLETA_CRONO")
hide_grid(ws)
set_widths(ws, [11, 14, 18, 14, 14, 8, 11, 9, 10, 11, 11, 10, 10, 12, 30])

ws["A1"] = "CRONOMETRAGEM DE CALIBRAÇÃO"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:O1")
ws["A2"] = "10 ciclos por peça Pareto · ritmo Westinghouse · tolerâncias 13% padrão"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:O2")

headers = [
    "DATA","MÁQUINA","PEÇA","OPERADOR","OBSERVADOR",
    "CICLO","T_CICLO_MIN","RITMO","FATOR_TOL","TN_MIN","TP_MIN",
    "SETUP_MIN","QTD_LOTE","TT_LOTE_MIN","OBS"
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

for r in range(5, 155):
    ws.cell(row=r, column=10, value=f'=IFERROR(G{r}*H{r},"")')
    ws.cell(row=r, column=11, value=f'=IFERROR(J{r}*(1+I{r}),"")')
    ws.cell(row=r, column=14, value=f'=IFERROR(L{r}+K{r}*M{r},"")')
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
borders_block(ws, 5, 1, 154, len(headers))
ws.freeze_panes = "A5"

# =====================================================================
# 6_COLETA_AMOSTRAGEM
# =====================================================================
ws = wb.create_sheet("6_COLETA_AMOSTRAGEM")
hide_grid(ws)
set_widths(ws, [11, 8, 11, 14, 14, 14, 30])

ws["A1"] = "AMOSTRAGEM DO TRABALHO"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:G1")
ws["A2"] = "~400 observações por máquina · IC 95% · erro 5%"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:G2")

headers = ["DATA","ROUND","HORA","MÁQUINA","STATUS","COD_PARADA","NOTA"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))
borders_block(ws, 5, 1, 504, len(headers))
for r in range(5, 505):
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)

dv = DataValidation(type="list",
    formula1='"PRODUZINDO,SETUP,PARADA,OCIOSO,AUSENTE"',
    allow_blank=True)
dv.add("E5:E504")
ws.add_data_validation(dv)
ws.freeze_panes = "A5"

# =====================================================================
# 7_TEMPOS_PADRAO
# =====================================================================
ws = wb.create_sheet("7_TEMPOS_PADRAO")
hide_grid(ws)
set_widths(ws, [14, 30, 14, 13, 14, 12, 14, 12, 12, 30])

ws["A1"] = "TEMPOS-PADRÃO OFICIAIS"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:J1")
ws["A2"] = "TP Crono × Tempo Apontado → Fator de Apontamento (FA)"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:J2")

headers = [
    "COD_PECA","DESCRIÇÃO","MÁQUINA","TP_CRONO_MIN","T_MED_APONT_MIN",
    "FA","TP_OFICIAL_MIN","SETUP_MIN","PCS/HORA","OBS"
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

for r in range(5, 55):
    ws.cell(row=r, column=6, value=f'=IFERROR(E{r}/D{r},"")')
    ws.cell(row=r, column=9, value=f'=IFERROR(60/G{r},"")')
    ws.cell(row=r, column=6).number_format = "0.00"
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
borders_block(ws, 5, 1, 54, len(headers))

# heatmap FA: <0,85 vermelho · 0,85-1,15 verde · >1,15 vermelho
ws.conditional_formatting.add("F5:F54", CellIsRule(
    operator="between", formula=["0.85", "1.15"], fill=fill("D5F5E3")))
ws.conditional_formatting.add("F5:F54", CellIsRule(
    operator="lessThan", formula=["0.85"], fill=fill("FADBD8")))
ws.conditional_formatting.add("F5:F54", CellIsRule(
    operator="greaterThan", formula=["1.15"], fill=fill("FADBD8")))

ws.freeze_panes = "A5"

# =====================================================================
# 8_CARGA_MAQUINA
# =====================================================================
ws = wb.create_sheet("8_CARGA_MAQUINA")
hide_grid(ws)
set_widths(ws, [14, 12, 12, 14, 14, 14, 14, 14, 30])

ws["A1"] = "CARGA-MÁQUINA"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:I1")
ws["A2"] = "Capacidade teórica × demanda real × ocupação · barra de cor por máquina"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:I2")

headers = [
    "MÁQUINA","DIAS_MES","H_DIA","CAP_TEÓRICA_H",
    "H_PRODUZIDA","H_SETUP","H_PARADA","OCUPAÇÃO_%","OBS"
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

for r in range(5, 25):
    ws.cell(row=r, column=4, value=f'=IFERROR(B{r}*C{r},"")')
    ws.cell(row=r, column=8, value=f'=IFERROR((E{r}+F{r})/D{r},"")')
    ws.cell(row=r, column=8).number_format = "0.0%"
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
borders_block(ws, 5, 1, 24, len(headers))

# data bar na ocupacao
ws.conditional_formatting.add("H5:H24",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=NAVY))
ws.freeze_panes = "A5"

# =====================================================================
# 9_PRODUTIVIDADE
# =====================================================================
ws = wb.create_sheet("9_PRODUTIVIDADE")
hide_grid(ws)
set_widths(ws, [14, 14, 14, 12, 14, 14, 14, 12, 30])

ws["A1"] = "PRODUTIVIDADE — MATRIZ O × M × P"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:I1")
ws["A2"] = "Eficiência = (QTD × TP_OFICIAL) ÷ T_REAL · heatmap por linha"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:I2")

headers = [
    "OPERADOR","MÁQUINA","PEÇA","QTD_PROD","TP_OFICIAL_MIN",
    "T_REAL_MIN","EFICIÊNCIA","FA_LOCAL","OBS"
]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

for r in range(5, 105):
    ws.cell(row=r, column=7, value=f'=IFERROR(D{r}*E{r}/F{r},"")')
    ws.cell(row=r, column=8, value=f'=IFERROR(F{r}/(D{r}*E{r}),"")')
    ws.cell(row=r, column=7).number_format = "0.0%"
    ws.cell(row=r, column=8).number_format = "0.00"
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
borders_block(ws, 5, 1, 104, len(headers))

# heatmap eficiencia
ws.conditional_formatting.add("G5:G104",
    ColorScaleRule(
        start_type="num", start_value=0.5, start_color="FADBD8",
        mid_type="num",   mid_value=0.85,  mid_color="FCF3CF",
        end_type="num",   end_value=1.0,   end_color="D5F5E3"
    ))
ws.freeze_panes = "A5"

# =====================================================================
# 10_PLANO_MELHORIA
# =====================================================================
ws = wb.create_sheet("10_PLANO_MELHORIA")
hide_grid(ws)
set_widths(ws, [6, 32, 16, 12, 12, 12, 14, 14, 14])

ws["A1"] = "PLANO DE MELHORIA"
ws["A1"].font = f(16, bold=True, color=NAVY)
ws.merge_cells("A1:I1")
ws["A2"] = "Ações priorizadas: PRIORIDADE = IMPACTO ÷ ESFORÇO (1–5 cada)"
ws["A2"].font = f(10, italic=True, color=GRAPH)
ws.merge_cells("A2:I2")

headers = ["ID","AÇÃO","ÁREA","IMPACTO","ESFORÇO","PRIORIDADE","RESPONSÁVEL","PRAZO","STATUS"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

for r in range(5, 35):
    ws.cell(row=r, column=1, value=f'=IF(B{r}<>"",ROW()-4,"")')
    ws.cell(row=r, column=6, value=f'=IFERROR(D{r}/E{r},"")')
    ws.cell(row=r, column=6).number_format = "0.00"
    if (r-5) % 2 == 1:
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = fill(LIGHT)
borders_block(ws, 5, 1, 34, len(headers))

dv_15 = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
dv_15.add("D5:E34")
ws.add_data_validation(dv_15)

dv_st = DataValidation(type="list",
    formula1='"ABERTA,EM ANDAMENTO,CONCLUÍDA,CANCELADA"', allow_blank=True)
dv_st.add("I5:I34")
ws.add_data_validation(dv_st)

ws.conditional_formatting.add("F5:F34",
    ColorScaleRule(
        start_type="num", start_value=0.4, start_color="FADBD8",
        mid_type="num",   mid_value=1,     mid_color="FCF3CF",
        end_type="num",   end_value=3,     end_color="D5F5E3"
    ))

ws.freeze_panes = "A5"

# =====================================================================
# salvar
# =====================================================================
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"OK: {OUT}")
