"""
processar_pdfs.py
Lê PDFs de apontamento da pasta fotos_apontamento/,
usa Claude Vision para extrair os dados e importa direto no CRONOANALISE.xlsx.

Uso:
    python processar_pdfs.py              # processa todos os PDFs novos
    python processar_pdfs.py nome.pdf     # processa um PDF específico

Requisitos:
    pip install anthropic pdf2image openpyxl pillow
    Variável de ambiente: ANTHROPIC_API_KEY
"""

import os
import sys
import csv
import shutil
import base64
import json
import re
from pathlib import Path
from datetime import datetime
from io import BytesIO

# ── dependências ──────────────────────────────────────────────────────────────
try:
    import anthropic
except ImportError:
    print("Instalando anthropic...")
    os.system("pip install anthropic --break-system-packages -q")
    import anthropic

try:
    from pdf2image import convert_from_path
except ImportError:
    print("Instalando pdf2image...")
    os.system("pip install pdf2image pillow --break-system-packages -q")
    from pdf2image import convert_from_path

try:
    import openpyxl
    from openpyxl.styles import Side, Border
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Side, Border

# ── configuração ──────────────────────────────────────────────────────────────
BASE        = Path(__file__).parent
PASTA_PDF   = BASE / "fotos_apontamento"
PASTA_PROC  = BASE / "fotos_apontamento" / "processados"
XLSX        = BASE / "CRONOANALISE.xlsx"
SHEET       = "4_APONTAMENTO_HIST"
FIRST_ROW   = 4

HEADERS = [
    "DATA","CLIENTE","ID_OP","ITEM","DESCRICAO_OP","N_PROG",
    "QTD_PROD","QTD_PED","H_INICIO","H_FIM","TOTAL_MIN","NC",
    "COD_PARADA","T_PARADA_MIN","OPERADOR","MAQUINA","AUDIT",
]

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMERIC_COLS = {3, 6, 7, 8, 11, 12, 14}   # colunas 1-indexed

# ── prompt para Claude Vision ─────────────────────────────────────────────────
PROMPT_EXTRACAO = """Você está lendo um formulário de APONTAMENTO DE PRODUÇÃO preenchido à mão.
Extraia TODAS as linhas de dados da tabela principal e retorne um JSON com a estrutura abaixo.

COLUNAS que você deve extrair (exatamente nesta ordem):
DATA, CLIENTE, ID_OP, ITEM, DESCRICAO_OP, N_PROG, QTD_PROD, QTD_PED,
H_INICIO, H_FIM, TOTAL_MIN, NC, COD_PARADA, T_PARADA_MIN, OPERADOR, MAQUINA, AUDIT

Regras importantes:
- DATA: formato DD/MM/AAAA (ex: 06/04/2026). Se só tiver dia/mês, complete com o ano do formulário.
- H_INICIO e H_FIM: formato HH:MM (ex: 07:30)
- QTD_PROD, QTD_PED, TOTAL_MIN, NC, T_PARADA_MIN: apenas números (sem texto)
- COD_PARADA: código tipo P1, P2, P3... Se vazio, deixe "".
- OPERADOR: nome como está escrito no formulário.
- MAQUINA: número ou código da máquina (campo no topo do formulário, geralmente "Máquina:").
- Se uma célula estiver vazia ou ilegível, use "".
- Ignore as linhas completamente em branco.
- Retorne SOMENTE o JSON, sem explicações.

Formato de retorno:
{
  "linhas": [
    {"DATA":"06/04/2026","CLIENTE":"MATAO","ID_OP":"31553","ITEM":"1.1","DESCRICAO_OP":"Fresar e alojar","N_PROG":"399","QTD_PROD":"70","QTD_PED":"50","H_INICIO":"07:30","H_FIM":"09:00","TOTAL_MIN":"90","NC":"","COD_PARADA":"","T_PARADA_MIN":"","OPERADOR":"Clayton","MAQUINA":"CNC01","AUDIT":""},
    ...
  ]
}
"""

def pdf_para_imagens_base64(pdf_path: Path) -> list[str]:
    """Converte PDF em lista de imagens base64 (uma por página)."""
    print(f"  Convertendo {pdf_path.name} para imagens...")
    imagens = convert_from_path(str(pdf_path), dpi=250)
    resultado = []
    for img in imagens:
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=90)
        b64 = base64.standard_b64encode(buf.getvalue()).decode("utf-8")
        resultado.append(b64)
    return resultado

def extrair_dados_claude(pdf_path: Path, api_key: str) -> list[dict]:
    """Usa Claude Vision para extrair dados de um PDF."""
    client = anthropic.Anthropic(api_key=api_key)
    imagens_b64 = pdf_para_imagens_base64(pdf_path)

    todas_linhas = []
    for i, img_b64 in enumerate(imagens_b64):
        print(f"  Processando página {i+1}/{len(imagens_b64)}...")
        msg = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": img_b64,
                        },
                    },
                    {"type": "text", "text": PROMPT_EXTRACAO},
                ],
            }],
        )

        texto = msg.content[0].text.strip()
        # Remove blocos ```json se existirem
        texto = re.sub(r"^```[a-z]*\n?", "", texto)
        texto = re.sub(r"\n?```$", "", texto)

        try:
            dados = json.loads(texto)
            linhas = dados.get("linhas", [])
            print(f"    → {len(linhas)} linhas extraídas")
            todas_linhas.extend(linhas)
        except json.JSONDecodeError as e:
            print(f"    AVISO: erro ao parsear JSON da página {i+1}: {e}")
            print(f"    Resposta recebida:\n{texto[:300]}")

    return todas_linhas

def to_num(v):
    """Converte string para número quando possível."""
    if not v:
        return None
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, OverflowError):
        return s

def importar_para_xlsx(linhas: list[dict], origem: str) -> int:
    """Insere linhas no CRONOANALISE.xlsx e retorna quantas foram inseridas."""
    if not linhas:
        return 0

    # backup
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    bkp = XLSX.with_name(f"CRONOANALISE_bkp_{stamp}.xlsx")
    shutil.copy2(XLSX, bkp)
    print(f"  Backup: {bkp.name}")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # acha primeira linha vazia
    first_empty = FIRST_ROW
    while ws.cell(row=first_empty, column=1).value not in (None, ""):
        first_empty += 1

    audit_val = f"OCR:{origem}/{stamp}"
    for i, linha in enumerate(linhas):
        for j, col in enumerate(HEADERS, 1):
            val = linha.get(col, "")
            # Coluna AUDIT: marca a origem
            if col == "AUDIT":
                val = audit_val
            # Converte colunas numéricas
            if j in NUMERIC_COLS and val != "":
                val = to_num(val)
            cell = ws.cell(row=first_empty + i, column=j, value=val if val != "" else None)
            cell.border = BORDER

    wb.save(XLSX)
    return len(linhas)

def mover_processado(pdf_path: Path):
    """Move PDF para subpasta 'processados/'."""
    PASTA_PROC.mkdir(exist_ok=True)
    dest = PASTA_PROC / pdf_path.name
    # Se já existir, adiciona timestamp
    if dest.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        dest = PASTA_PROC / f"{pdf_path.stem}_{stamp}{pdf_path.suffix}"
    shutil.move(str(pdf_path), str(dest))
    print(f"  Movido para: processados/{dest.name}")

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    # Verifica API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("ERRO: variável ANTHROPIC_API_KEY não definida.")
        print("  Defina com:  set ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    # Verifica Excel
    if not XLSX.exists():
        print(f"ERRO: {XLSX} não encontrado.")
        sys.exit(1)

    # Quais PDFs processar?
    if len(sys.argv) > 1:
        pdfs = []
        for arg in sys.argv[1:]:
            p = Path(arg)
            if not p.is_absolute():
                p = PASTA_PDF / p
            if not p.exists():
                print(f"AVISO: {p} não encontrado, pulando.")
            else:
                pdfs.append(p)
    else:
        pdfs = sorted(PASTA_PDF.glob("*.pdf"))

    if not pdfs:
        print("Nenhum PDF encontrado em fotos_apontamento/")
        sys.exit(0)

    total_linhas = 0
    for pdf in pdfs:
        print(f"\n{'='*50}")
        print(f"Processando: {pdf.name}")
        try:
            linhas = extrair_dados_claude(pdf, api_key)
            if linhas:
                n = importar_para_xlsx(linhas, pdf.stem)
                print(f"  ✓ {n} linhas importadas para {XLSX.name}")
                total_linhas += n
                mover_processado(pdf)
            else:
                print(f"  AVISO: nenhuma linha extraída de {pdf.name}")
        except Exception as e:
            print(f"  ERRO ao processar {pdf.name}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*50}")
    print(f"Concluído! {total_linhas} linhas importadas no total.")
    print(f"Abra o Power BI e clique em 'Atualizar' para ver os novos dados.")

if __name__ == "__main__":
    main()
